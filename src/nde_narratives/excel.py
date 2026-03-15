from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .config import PathsConfig, StudyConfig
from .constants import ANNOTATION_SHEET, COLUMN_MAP_SHEET, INSTRUCTIONS_SHEET, MAPPING_SHEET, SAMPLED_PRIVATE_SHEET


def _excel_column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def add_excel_validation_and_formatting(xlsx_path: Path, study: StudyConfig) -> None:
    workbook = load_workbook(xlsx_path)
    worksheet = workbook[ANNOTATION_SHEET]

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    worksheet.freeze_panes = "A2"

    for column_cells in worksheet.columns:
        header = str(column_cells[0].value)
        column_letter = column_cells[0].column_letter
        if "Narrative" in header:
            worksheet.column_dimensions[column_letter].width = 50
        elif "Tone" in header:
            worksheet.column_dimensions[column_letter].width = 18
        elif header == "Participant Code":
            worksheet.column_dimensions[column_letter].width = 20
        else:
            worksheet.column_dimensions[column_letter].width = 24

    columns = list(pd.read_excel(xlsx_path, sheet_name=ANNOTATION_SHEET, nrows=0).columns)
    column_to_index = {name: index + 1 for index, name in enumerate(columns)}
    max_row = worksheet.max_row

    tone_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(study.tone_labels)}"',
        allow_blank=True,
    )
    binary_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(study.binary_labels)}"',
        allow_blank=True,
    )
    worksheet.add_data_validation(tone_validation)
    worksheet.add_data_validation(binary_validation)

    for section_name in study.section_order:
        section = study.sections[section_name]
        tone_letter = _excel_column_letter(column_to_index[section.tone_visible_label])
        tone_validation.add(f"{tone_letter}2:{tone_letter}{max_row}")
        for visible_label in section.binary_labels.values():
            binary_letter = _excel_column_letter(column_to_index[visible_label])
            binary_validation.add(f"{binary_letter}2:{binary_letter}{max_row}")

    instructions = workbook[INSTRUCTIONS_SHEET] if INSTRUCTIONS_SHEET in workbook.sheetnames else workbook.create_sheet(INSTRUCTIONS_SHEET)
    instructions.delete_rows(1, instructions.max_row)
    rows = [
        ["Field", "Instruction"],
        ["Participant Code", "Anonymous identifier. Do not modify."],
        ["Context Narrative", "Read the text and assign only the Context Tone."],
        ["Experience Narrative", "Read the text and assign Experience Tone plus the related yes/no elements."],
        ["Aftereffects Narrative", "Read the text and assign Aftereffects Tone plus the related yes/no elements."],
        ["Tone labels", f"Allowed values: {', '.join(study.tone_labels)}"],
        ["Binary labels", f"Allowed values: {', '.join(study.binary_labels)}"],
        ["General rule", "Code only what is explicit in the text. If unclear, use no for binary elements."],
    ]
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            instructions.cell(row=row_index, column=column_index, value=value)

    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 95

    workbook.save(xlsx_path)


def write_annotation_outputs(
    annotation_df: pd.DataFrame,
    mapping_df: pd.DataFrame,
    column_map_df: pd.DataFrame,
    sampled_private_df: pd.DataFrame,
    study: StudyConfig,
    paths: PathsConfig,
) -> dict[str, str]:
    paths.annotation_output_dir.mkdir(parents=True, exist_ok=True)
    paths.sampled_private_workbook.parent.mkdir(parents=True, exist_ok=True)

    annotator_path = paths.annotation_output_dir / study.outputs["annotator_filename"]
    mapping_path = paths.annotation_output_dir / study.outputs["mapping_filename"]
    column_map_path = paths.annotation_output_dir / study.outputs["column_map_filename"]

    with pd.ExcelWriter(annotator_path, engine="openpyxl") as writer:
        annotation_df.to_excel(writer, sheet_name=ANNOTATION_SHEET, index=False)

    add_excel_validation_and_formatting(annotator_path, study)

    with pd.ExcelWriter(mapping_path, engine="openpyxl") as writer:
        mapping_df.to_excel(writer, sheet_name=MAPPING_SHEET, index=False)
        sampled_private_df.to_excel(writer, sheet_name=SAMPLED_PRIVATE_SHEET, index=False)

    with pd.ExcelWriter(column_map_path, engine="openpyxl") as writer:
        column_map_df.to_excel(writer, sheet_name=COLUMN_MAP_SHEET, index=False)

    return {
        "annotator_file": str(annotator_path),
        "mapping_file": str(mapping_path),
        "column_map_file": str(column_map_path),
    }
