from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nde_narratives.config import load_study_config
from nde_narratives.constants import ANNOTATION_SHEET, PARTICIPANT_CODE_HEADER


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"


def run_cli(*args: str) -> subprocess.CompletedProcess[str]:
    command = [sys.executable, "-m", "nde_narratives.cli", *args]
    return subprocess.run(command, cwd=ROOT, check=False, capture_output=True, text=True)


def make_paths_config(tmp_path: Path, survey_csv: Path) -> Path:
    annotation_dir = tmp_path / "annotation_outputs"
    llm_dir = tmp_path / "llm_batches"
    evaluation_dir = tmp_path / "evaluation_outputs"
    llm_output_dir = tmp_path / "llm_outputs"

    for directory in (annotation_dir, llm_dir, evaluation_dir, llm_output_dir):
        directory.mkdir(parents=True, exist_ok=True)

    content = f"""[paths]
survey_csv = \"{survey_csv.as_posix()}\"
annotation_output_dir = \"{annotation_dir.as_posix()}\"
llm_batch_dir = \"{llm_dir.as_posix()}\"
evaluation_output_dir = \"{evaluation_dir.as_posix()}\"
sampled_private_workbook = \"{(annotation_dir / 'nde_annotation_mapping_private.xlsx').as_posix()}\"
human_annotation_workbook = \"{(annotation_dir / 'nde_annotation_sample.xlsx').as_posix()}\"
llm_predictions_path = \"{(llm_output_dir / 'nde_predictions.jsonl').as_posix()}\"
"""
    path = tmp_path / "paths.local.toml"
    path.write_text(content, encoding="utf-8")
    return path


def test_validate_config_passes_with_synthetic_fixture(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    result = run_cli("validate-config", "--study-config", str(study_config), "--paths-config", str(paths_config))

    assert result.returncode == 0, result.stderr
    assert "Configuration valid." in result.stdout


def test_build_annotation_sample_creates_workbooks(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))

    assert result.returncode == 0, result.stderr
    annotation_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    column_map_workbook = tmp_path / "annotation_outputs" / "nde_annotation_column_mapping_private.xlsx"

    assert annotation_workbook.exists()
    assert mapping_workbook.exists()
    assert column_map_workbook.exists()

    workbook = load_workbook(annotation_workbook)
    assert set(workbook.sheetnames) == {"annotation", "instructions"}
    worksheet = workbook["annotation"]
    headers = [cell.value for cell in worksheet[1]]
    assert PARTICIPANT_CODE_HEADER in headers
    assert "Context Narrative" in headers
    assert len(worksheet.data_validations.dataValidation) == 2

    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private")
    assert len(sampled_private) == 3
    assert "participant_code" in sampled_private.columns


def test_build_llm_batch_creates_three_section_files(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    result = run_cli("build-llm-batch", "--study-config", str(study_config), "--paths-config", str(paths_config))

    assert result.returncode == 0, result.stderr

    for section in ("context", "experience", "aftereffects"):
        batch_path = tmp_path / "llm_batches" / f"{section}_batch.jsonl"
        assert batch_path.exists()
        records = [json.loads(line) for line in batch_path.read_text(encoding="utf-8").splitlines() if line]
        assert len(records) == 3
        assert {record["section"] for record in records} == {section}
        assert all(record["response_schema"]["title"].startswith("nde_") for record in records)


def test_sentiment_sensitivity_generates_scores_figures_and_report(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)
    output_dir = tmp_path / "sentiment_outputs"

    result = run_cli(
        "sentiment-sensitivity",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--output-dir",
        str(output_dir),
    )

    assert result.returncode == 0, result.stderr
    scores_path = output_dir / "vader_sentiment_scores.csv"
    report_path = output_dir / "vader_report.md"
    assert scores_path.exists()
    assert report_path.exists()
    for section in ("context", "experience", "aftereffects"):
        assert (output_dir / "figures" / f"{section}_distribution.png").exists()

    scores = pd.read_csv(scores_path)
    assert len(scores) == 9
    assert set(scores["section"]) == {"context", "experience", "aftereffects"}
    assert set(scores["vader_label"]).issubset({"positive", "negative", "mixed"})
    assert "text" not in scores.columns

    report_text = report_path.read_text(encoding="utf-8")
    assert "## Methodology" in report_text
    assert "## Results" in report_text

    all_records_result = run_cli(
        "sentiment-sensitivity",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--output-dir",
        str(tmp_path / "sentiment_outputs_all"),
        "--all-records",
    )
    assert all_records_result.returncode == 0, all_records_result.stderr
    all_scores = pd.read_csv(tmp_path / "sentiment_outputs_all" / "vader_sentiment_scores.csv")
    assert len(all_scores) == 12


    include_text_result = run_cli(
        "sentiment-sensitivity",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--output-dir",
        str(tmp_path / "sentiment_outputs_text"),
        "--include-text",
    )
    assert include_text_result.returncode == 0, include_text_result.stderr
    include_text_scores = pd.read_csv(tmp_path / "sentiment_outputs_text" / "vader_sentiment_scores.csv")
    assert "text" in include_text_scores.columns


def populate_human_annotation_workbook(annotation_workbook: Path, mapping_workbook: Path, study_config: Path) -> None:
    study = load_study_config(study_config)
    annotation_values = pd.read_csv(FIXTURES / "manual_annotations_fixture.csv").set_index("response_id")
    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private").set_index("participant_code")
    internal_to_visible = study.internal_to_visible_annotation_columns()

    workbook = load_workbook(annotation_workbook)
    worksheet = workbook[ANNOTATION_SHEET]
    headers = {cell.value: index for index, cell in enumerate(worksheet[1], start=1)}
    code_column = headers[PARTICIPANT_CODE_HEADER]

    for row_index in range(2, worksheet.max_row + 1):
        participant_code = worksheet.cell(row=row_index, column=code_column).value
        response_id = sampled_private.loc[participant_code, study.id_column]
        values = annotation_values.loc[response_id]
        for column in study.annotation_internal_columns():
            visible_column = internal_to_visible[column]
            worksheet.cell(row=row_index, column=headers[visible_column], value=values[column])

    workbook.save(annotation_workbook)


def write_llm_predictions_fixture(mapping_workbook: Path, study_config: Path, output_path: Path) -> None:
    study = load_study_config(study_config)
    predictions = pd.read_csv(FIXTURES / "llm_predictions_fixture.csv").set_index("response_id")
    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private").set_index("participant_code")

    with output_path.open("w", encoding="utf-8") as handle:
        for participant_code, row in sampled_private.iterrows():
            response_id = row[study.id_column]
            prediction_row = predictions.loc[response_id]
            for section in study.section_order:
                payload = {"participant_code": participant_code, "section": section}
                section_config = study.sections[section]
                payload[section_config.tone_internal_column] = prediction_row[section_config.tone_internal_column]
                for column in section_config.binary_labels:
                    payload[column] = prediction_row[column]
                handle.write(json.dumps(payload))
                handle.write("\n")


def test_evaluate_reuses_existing_vader_scores_and_reports_vader_metrics(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    annotation_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    prediction_path = tmp_path / "llm_outputs" / "nde_predictions.jsonl"
    vader_dir = tmp_path / "evaluation_outputs" / "vader_sentiment"

    populate_human_annotation_workbook(annotation_workbook, mapping_workbook, study_config)
    write_llm_predictions_fixture(mapping_workbook, study_config, prediction_path)

    vader_result = run_cli(
        "sentiment-sensitivity",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--output-dir",
        str(vader_dir),
    )
    assert vader_result.returncode == 0, vader_result.stderr

    eval_result = run_cli(
        "evaluate",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--vader-scores",
        str(vader_dir / "vader_sentiment_scores.csv"),
    )
    assert eval_result.returncode == 0, eval_result.stderr

    metrics = pd.read_csv(tmp_path / "evaluation_outputs" / "evaluation_metrics.csv")
    assert set(metrics["comparison"]).issuperset({"human_vs_vader", "llm_vs_vader"})
    assert len(metrics[metrics["comparison"] == "human_vs_vader"]) == 3

    workbook = load_workbook(annotation_workbook)
    worksheet = workbook[ANNOTATION_SHEET]
    headers = {cell.value: index for index, cell in enumerate(worksheet[1], start=1)}
    worksheet.cell(row=2, column=headers["Context Tone"], value="invalid")
    workbook.save(annotation_workbook)

    invalid_result = run_cli("evaluate", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert invalid_result.returncode == 1
    assert "invalid labels" in invalid_result.stderr


def test_evaluate_generates_sample_vader_scores_when_global_file_is_missing(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    annotation_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    prediction_path = tmp_path / "llm_outputs" / "nde_predictions.jsonl"

    populate_human_annotation_workbook(annotation_workbook, mapping_workbook, study_config)
    write_llm_predictions_fixture(mapping_workbook, study_config, prediction_path)

    eval_result = run_cli("evaluate", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert eval_result.returncode == 0, eval_result.stderr

    sample_scores_path = tmp_path / "evaluation_outputs" / "vader_sentiment_sample" / "vader_sentiment_scores.csv"
    assert sample_scores_path.exists()
    metrics = pd.read_csv(tmp_path / "evaluation_outputs" / "evaluation_metrics.csv")
    assert len(metrics[metrics["comparison"] == "human_vs_vader"]) == 3
