from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nde_narratives.constants import ANNOTATION_SHEET

from tests.cli_helpers import (
    FIXTURES,
    copy_completed_annotation_to_human_dir,
    make_paths_config,
    populate_human_annotation_workbook,
    run_cli,
    write_human_manifest,
    write_llm_manifest,
    write_llm_predictions_fixture,
)


def _annotation_headers(worksheet) -> dict[str, int]:
    return {cell.value: index for index, cell in enumerate(worksheet[1], start=1)}


def _set_annotation_value(annotation_workbook: Path, row_index: int, header: str, value: str) -> None:
    workbook = load_workbook(annotation_workbook)
    worksheet = workbook[ANNOTATION_SHEET]
    headers = _annotation_headers(worksheet)
    worksheet.cell(row=row_index, column=headers[header], value=value)
    workbook.save(annotation_workbook)


def _add_extra_annotation_column(annotation_workbook: Path, header: str, value: str) -> None:
    workbook = load_workbook(annotation_workbook)
    worksheet = workbook[ANNOTATION_SHEET]
    new_column_index = worksheet.max_column + 1
    worksheet.cell(row=1, column=new_column_index, value=header)
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=new_column_index, value=value)
    workbook.save(annotation_workbook)


def _prepare_human_artifact(source_workbook: Path, human_root: Path, annotator_id: str) -> Path:
    target_dir = human_root / annotator_id
    target_dir.mkdir(parents=True, exist_ok=True)
    target = copy_completed_annotation_to_human_dir(source_workbook, target_dir / f"{annotator_id}.xlsx")
    write_human_manifest(target_dir, annotator_id)
    return target


def _prepare_llm_artifact(
    mapping_workbook: Path,
    study_config: Path,
    llm_root: Path,
    experiment_id: str,
    run_id: str = "run-01",
    prompt_variant: str = "baseline",
) -> Path:
    target_dir = llm_root / f"{experiment_id}_{run_id}"
    target_dir.mkdir(parents=True, exist_ok=True)
    prediction_path = target_dir / "predictions.jsonl"
    write_llm_predictions_fixture(mapping_workbook, study_config, prediction_path)
    write_llm_manifest(target_dir, experiment_id=experiment_id, prompt_variant=prompt_variant, run_id=run_id, model_variant="mock-model")
    return prediction_path


def test_evaluate_uses_majority_reference_and_reports_artifacts(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    source_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    human_root = tmp_path / "human_annotations"
    llm_root = tmp_path / "llm_outputs"

    populate_human_annotation_workbook(source_workbook, mapping_workbook, study_config)
    annotator_a = _prepare_human_artifact(source_workbook, human_root, "ann_a")
    annotator_b = _prepare_human_artifact(source_workbook, human_root, "ann_b")
    workbook = load_workbook(annotator_b)
    worksheet = workbook[ANNOTATION_SHEET]
    headers = _annotation_headers(worksheet)
    current_value = worksheet.cell(row=2, column=headers["Context Tone"]).value
    replacement = next(label for label in ("positive", "negative", "mixed") if label != current_value)
    _set_annotation_value(annotator_b, 2, "Context Tone", replacement)
    _add_extra_annotation_column(annotator_b, "extra_reviewer_note", "not used in metrics")

    invalid_human_dir = human_root / "broken"
    invalid_human_dir.mkdir(parents=True, exist_ok=True)
    (invalid_human_dir / "broken.xlsx").write_text("not an xlsx", encoding="utf-8")
    write_human_manifest(invalid_human_dir, "broken")

    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-alpha")
    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-beta", run_id="run-02")
    invalid_llm_dir = llm_root / "bad-exp"
    invalid_llm_dir.mkdir(parents=True, exist_ok=True)
    (invalid_llm_dir / "predictions.jsonl").write_text('{"participant_code": "ANN_0001"}\n', encoding="utf-8")
    write_llm_manifest(invalid_llm_dir, experiment_id="bad-exp")

    eval_result = run_cli("evaluate", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert eval_result.returncode == 0, eval_result.stderr

    metrics = pd.read_csv(tmp_path / "evaluation_outputs" / "evaluation_metrics.csv")
    summary = json.loads((tmp_path / "evaluation_outputs" / "evaluation_summary.json").read_text(encoding="utf-8"))
    report_text = (tmp_path / "evaluation_outputs" / "alignment_report.md").read_text(encoding="utf-8")
    questionnaire_report_text = (tmp_path / "evaluation_outputs" / "alignment_report_questionnaire.md").read_text(encoding="utf-8")
    reference_df = pd.read_csv(tmp_path / "evaluation_outputs" / "human_reference_majority.csv")
    family_metrics = pd.read_csv(tmp_path / "evaluation_outputs" / "alignment_family_metrics.csv")
    human_pairwise = pd.read_csv(tmp_path / "evaluation_outputs" / "human_agreement_pairwise.csv")
    llm_manifest = json.loads((tmp_path / "evaluation_outputs" / "llm_artifacts_manifest.json").read_text(encoding="utf-8"))
    human_manifest = json.loads((tmp_path / "evaluation_outputs" / "human_artifacts_manifest.json").read_text(encoding="utf-8"))

    context_vader = metrics[(metrics["comparison"] == "human_reference_vs_vader") & (metrics["field"] == "context_tone")].iloc[0]
    llm_rows = metrics[metrics["comparison"].str.startswith("human_reference_vs_llm:")]

    assert int(context_vader["n"]) == 2
    assert len(reference_df) == 3
    assert summary["coverage"]["n_valid_human_artifacts"] == 2
    assert summary["coverage"]["n_rejected_human_artifacts"] == 1
    assert summary["coverage"]["n_valid_llm_artifacts"] == 2
    assert summary["coverage"]["n_rejected_llm_artifacts"] == 1
    assert summary["adjudication"]["n_unresolved_field_participant_pairs"] >= 1
    assert not human_pairwise.empty
    assert set(llm_manifest.keys()) == {"accepted", "rejected"}
    assert len(llm_manifest["accepted"]) == 2
    assert len(llm_manifest["rejected"]) == 1
    assert len(human_manifest["accepted"]) == 2
    assert len(human_manifest["rejected"]) == 1
    assert set(llm_rows["artifact_id"]) == {"exp_alpha__run-01", "exp_beta__run-02"}
    assert {"questionnaire_vs_vader", "questionnaire_vs_llm:exp_alpha__run-01", "questionnaire_vs_llm:exp_beta__run-02"}.issubset(set(metrics["comparison"]))
    assert {"vader_vs_llm:exp_alpha__run-01", "vader_vs_llm:exp_beta__run-02", "llm_vs_llm:exp_alpha__run-01__vs__exp_beta__run-02"}.issubset(set(metrics["comparison"]))
    questionnaire_vader_fields = set(metrics.loc[metrics["comparison"] == "questionnaire_vs_vader", "field"])
    assert questionnaire_vader_fields == {"experience_tone"}
    tone_label_columns = {"f1_label_positive", "f1_label_negative", "f1_label_mixed", "f1_label_neutral"}
    assert tone_label_columns.issubset(set(metrics.columns))
    assert "experience_tone_label_f1" in summary
    assert summary["experience_tone_label_f1"]["labels"] == sorted(tone_label_columns)
    assert len(summary["experience_tone_label_f1"]["rows"]) >= 1
    assert "majority vote" in report_text.lower()
    assert "rejected human artifacts" in report_text.lower()
    assert "accepted experiments evaluated against the human majority reference" in report_text.lower()
    assert "paper figures (lead)" in report_text.lower()
    assert "hypothesis verdict" in report_text.lower()
    assert "global evidence (macro f1)" in report_text.lower()
    assert "family-level support (macro f1)" in report_text.lower()
    assert "human_family_summary.png" in report_text
    assert "human_tone_summary.png" in report_text
    assert "human_tone_alignment.png" in report_text
    assert "extended figures" in report_text.lower()
    assert "human_comparison_summary.png" in report_text
    assert report_text.index("## Paper Figures (Lead)") < report_text.index("## Hypothesis Verdict")
    assert report_text.index("## Hypothesis Verdict") < report_text.index("## General Results")
    assert "general results" in report_text.lower()
    assert "family-level results" in report_text.lower()
    assert "item-level detail" in report_text.lower()
    assert "family-level summary" in report_text.lower()
    assert "questionnaire vs automated" in questionnaire_report_text.lower()
    assert "contradiction-focused qualitative analysis" in questionnaire_report_text.lower()
    assert "curated contradictory evidence examples" in questionnaire_report_text.lower()
    assert "questionnaire_contradiction_overview.png" in questionnaire_report_text
    assert "questionnaire_contradiction_unigram_wordcloud.png" in questionnaire_report_text
    assert "questionnaire_tone_confusion_matrix.png" in questionnaire_report_text
    assert "questionnaire_tone_summary.png" not in questionnaire_report_text
    assert "tone label confusion and per-label f1" in questionnaire_report_text.lower()
    assert "neutral focus in experience tone" not in questionnaire_report_text.lower()
    detailed_heatmaps_index = questionnaire_report_text.index("## Item-Level Structure (Extended Figures)")
    contradiction_index = questionnaire_report_text.index("## Contradiction-Focused Qualitative Analysis")
    assert detailed_heatmaps_index < contradiction_index
    assert set(family_metrics["family"]) >= {"tone", "m8", "m9"}
    assert "families are operationalized as tone, nde-c, and nde-mcq" in report_text.lower()
    assert "#### NDE-C" in report_text
    assert "#### NDE-MCQ" in report_text
    assert "#### M8" not in report_text
    assert "#### M9" not in report_text
    assert "questionnaire_extraction_item_scatter.png" in questionnaire_report_text
    assert "questionnaire_nde_c_macro_f1_heatmap.png" not in questionnaire_report_text
    assert "questionnaire_nde_mcq_macro_f1_heatmap.png" not in questionnaire_report_text
    assert "questionnaire_m8_accuracy_heatmap.png" not in questionnaire_report_text
    assert "questionnaire_m9_accuracy_heatmap.png" not in questionnaire_report_text
    assert "<details open>" in report_text
    assert (tmp_path / "evaluation_outputs" / "experiments" / "exp_alpha__run-01" / "evaluation_metrics.csv").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "human_family_summary.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "human_nde_c_macro_f1_heatmap.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "human_nde_mcq_macro_f1_heatmap.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_extraction_item_scatter.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_family_tradeoff_map.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_tone_confusion_matrix.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_contradiction_overview.png").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_contradiction_unigram_wordcloud.png").exists()
    assert (tmp_path / "evaluation_outputs" / "questionnaire_contradictions.csv").exists()
    assert "questionnaire_contradictions" in summary
    assert "questionnaire_family_tradeoff_map.png" in questionnaire_report_text


def test_evaluate_can_export_pdf_figures(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    source_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    human_root = tmp_path / "human_annotations"
    llm_root = tmp_path / "llm_outputs"

    populate_human_annotation_workbook(source_workbook, mapping_workbook, study_config)
    _prepare_human_artifact(source_workbook, human_root, "ann_a")
    _prepare_human_artifact(source_workbook, human_root, "ann_b")
    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-alpha")

    eval_result = run_cli(
        "evaluate",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--export-figures-pdf",
        "--figure-dpi",
        "360",
    )
    assert eval_result.returncode == 0, eval_result.stderr
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "human_comparison_summary.pdf").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "human_family_summary.pdf").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_family_tradeoff_map.pdf").exists()
    assert (tmp_path / "evaluation_outputs" / "figures" / "alignment" / "questionnaire_contradiction_overview.pdf").exists()


def test_evaluate_can_filter_to_selected_experiment_and_annotator(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    source_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    human_root = tmp_path / "human_annotations"
    llm_root = tmp_path / "llm_outputs"

    populate_human_annotation_workbook(source_workbook, mapping_workbook, study_config)
    _prepare_human_artifact(source_workbook, human_root, "keep_me")
    _prepare_human_artifact(source_workbook, human_root, "drop_me")

    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-one", run_id="run-01")
    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-two", run_id="run-02")

    eval_result = run_cli(
        "evaluate",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--annotator-id",
        "keep_me",
        "--experiment-id",
        "exp-two",
    )
    assert eval_result.returncode == 0, eval_result.stderr

    metrics = pd.read_csv(tmp_path / "evaluation_outputs" / "evaluation_metrics.csv")
    summary = json.loads((tmp_path / "evaluation_outputs" / "evaluation_summary.json").read_text(encoding="utf-8"))

    llm_rows = metrics[metrics["comparison"].str.startswith("human_reference_vs_llm:")]
    assert set(llm_rows["artifact_id"]) == {"exp_two__run-02"}
    assert summary["coverage"]["n_valid_human_artifacts"] == 1
    assert summary["coverage"]["n_valid_llm_artifacts"] == 1


def test_evaluate_can_filter_to_selected_prompt_variant_and_custom_output_dir(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    build_result = run_cli("build-annotation-sample", "--study-config", str(study_config), "--paths-config", str(paths_config))
    assert build_result.returncode == 0, build_result.stderr

    source_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    human_root = tmp_path / "human_annotations"
    llm_root = tmp_path / "llm_outputs"

    populate_human_annotation_workbook(source_workbook, mapping_workbook, study_config)
    _prepare_human_artifact(source_workbook, human_root, "ann_a")

    _prepare_llm_artifact(mapping_workbook, study_config, llm_root, "exp-baseline", run_id="run-01", prompt_variant="baseline")
    _prepare_llm_artifact(
        mapping_workbook,
        study_config,
        llm_root,
        "exp-smaj",
        run_id="run-01",
        prompt_variant="sentence_majority_v1",
    )

    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private")
    vader_rows: list[dict[str, object]] = []
    for _, row in sampled_private.iterrows():
        for section_name in ("context", "experience", "aftereffects"):
            vader_rows.append(
                {
                    "response_id": row["response_id"],
                    "participant_code": row["participant_code"],
                    "section": section_name,
                    "vader_label": "positive",
                    "compound": 0.7,
                    "neg": 0.0,
                    "neu": 0.3,
                    "pos": 0.7,
                }
            )
    vader_scores_path = tmp_path / "vader_scores_fixture.csv"
    pd.DataFrame(vader_rows).to_csv(vader_scores_path, index=False)

    custom_output_dir = tmp_path / "evaluation_outputs_sentence_majority"
    eval_result = run_cli(
        "evaluate",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--prompt-variant",
        "sentence_majority_v1",
        "--vader-scores",
        str(vader_scores_path),
        "--output-dir",
        str(custom_output_dir),
    )
    assert eval_result.returncode == 0, eval_result.stderr

    metrics = pd.read_csv(custom_output_dir / "evaluation_metrics.csv")
    summary = json.loads((custom_output_dir / "evaluation_summary.json").read_text(encoding="utf-8"))
    llm_manifest = json.loads((custom_output_dir / "llm_artifacts_manifest.json").read_text(encoding="utf-8"))

    llm_rows = metrics[metrics["comparison"].str.startswith("human_reference_vs_llm:")]
    assert set(llm_rows["artifact_id"]) == {"exp_smaj__run-01"}
    assert summary["coverage"]["n_valid_llm_artifacts"] == 1
    assert len(llm_manifest["accepted"]) == 1
    assert llm_manifest["accepted"][0]["prompt_variant"] == "sentence_majority_v1"
    assert (custom_output_dir / "alignment_report.md").exists()
    assert not (tmp_path / "evaluation_outputs" / "evaluation_metrics.csv").exists()
