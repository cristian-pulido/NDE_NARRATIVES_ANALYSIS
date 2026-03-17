from __future__ import annotations

import json
import os
import subprocess
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nde_narratives.config import load_study_config
from nde_narratives.constants import ANNOTATION_SHEET, PARTICIPANT_CODE_HEADER


ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "tests" / "fixtures"


def run_cli(*args: str, env: dict[str, str] | None = None) -> subprocess.CompletedProcess[str]:
    command = [sys.executable, "-m", "nde_narratives.cli", *args]
    merged_env = os.environ.copy()
    if env:
        merged_env.update(env)
    return subprocess.run(command, cwd=ROOT, check=False, capture_output=True, text=True, env=merged_env)


def make_paths_config(tmp_path: Path, survey_csv: Path, llm_block: str | None = None) -> Path:
    annotation_dir = tmp_path / "annotation_outputs"
    human_dir = tmp_path / "human_annotations"
    llm_dir = tmp_path / "llm_batches"
    evaluation_dir = tmp_path / "evaluation_outputs"
    llm_output_dir = tmp_path / "llm_outputs"
    prompt_variants_dir = tmp_path / "prompt_variants"

    for directory in (annotation_dir, human_dir, llm_dir, evaluation_dir, llm_output_dir, prompt_variants_dir):
        directory.mkdir(parents=True, exist_ok=True)

    content = f'''[paths]
survey_csv = "{survey_csv.as_posix()}"
annotation_output_dir = "{annotation_dir.as_posix()}"
human_annotations_dir = "{human_dir.as_posix()}"
llm_batch_dir = "{llm_dir.as_posix()}"
llm_results_dir = "{llm_output_dir.as_posix()}"
evaluation_output_dir = "{evaluation_dir.as_posix()}"
prompt_variants_dir = "{prompt_variants_dir.as_posix()}"
sampled_private_workbook = "{(annotation_dir / 'nde_annotation_mapping_private.xlsx').as_posix()}"
human_annotation_workbook = "{(human_dir / 'nde_annotation_sample.xlsx').as_posix()}"
llm_predictions_path = "{(llm_output_dir / 'nde_predictions.jsonl').as_posix()}"
'''
    if llm_block:
        content += "\n" + llm_block.strip() + "\n"
    path = tmp_path / "paths.local.toml"
    path.write_text(content, encoding="utf-8")
    return path


def populate_human_annotation_workbook(annotation_workbook: Path, mapping_workbook: Path, study_config: Path) -> None:
    study = load_study_config(study_config)
    annotation_values = pd.read_csv(FIXTURES / "manual_annotations_fixture.csv").set_index("response_id")
    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private").set_index("participant_code")
    internal_to_visible = study.internal_to_visible_annotation_columns()

    workbook = load_workbook(annotation_workbook)
    worksheet = workbook[ANNOTATION_SHEET]
    headers = {cell.value: index for index, cell in enumerate(worksheet[1], start=1)}
    code_column = headers[PARTICIPANT_CODE_HEADER]

    for row_index in range(2, worksheet.max_row + 1):
        participant_code = worksheet.cell(row=row_index, column=code_column).value
        response_id = sampled_private.loc[participant_code, study.id_column]
        values = annotation_values.loc[response_id]
        for column in study.annotation_internal_columns():
            visible_column = internal_to_visible[column]
            worksheet.cell(row=row_index, column=headers[visible_column], value=values[column])

    workbook.save(annotation_workbook)


def copy_completed_annotation_to_human_dir(annotation_workbook: Path, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(annotation_workbook.read_bytes())
    return destination


def write_human_manifest(target_dir: Path, annotator_id: str) -> Path:
    path = target_dir / "manifest.json"
    path.write_text(json.dumps({"annotator_id": annotator_id}, indent=2), encoding="utf-8")
    return path


def write_llm_predictions_fixture(mapping_workbook: Path, study_config: Path, output_path: Path) -> None:
    study = load_study_config(study_config)
    predictions = pd.read_csv(FIXTURES / "llm_predictions_fixture.csv").set_index("response_id")
    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private").set_index("participant_code")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as handle:
        for participant_code, row in sampled_private.iterrows():
            response_id = row[study.id_column]
            prediction_row = predictions.loc[response_id]
            for section in study.section_order:
                payload = {"participant_code": participant_code, "section": section}
                section_config = study.sections[section]
                payload[section_config.tone_internal_column] = prediction_row[section_config.tone_internal_column]
                for column in section_config.binary_labels:
                    payload[column] = prediction_row[column]
                handle.write(json.dumps(payload))
                handle.write("\n")


def write_llm_manifest(target_dir: Path, experiment_id: str, prompt_variant: str | None = None, run_id: str | None = None, model_variant: str | None = None) -> Path:
    payload = {"experiment_id": experiment_id}
    if prompt_variant is not None:
        payload["prompt_variant"] = prompt_variant
    if run_id is not None:
        payload["run_id"] = run_id
    if model_variant is not None:
        payload["model_variant"] = model_variant
    path = target_dir / "manifest.json"
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return path
