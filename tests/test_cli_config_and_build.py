from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from nde_narratives.config import load_benchmark_config
from nde_narratives.constants import PARTICIPANT_CODE_HEADER

from tests.cli_helpers import FIXTURES, make_paths_config, run_cli


def test_validate_config_passes_with_synthetic_fixture(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    result = run_cli(
        "validate-config",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    assert "Configuration valid." in result.stdout


def test_top_level_help_is_more_descriptive() -> None:
    result = run_cli("--help")

    assert result.returncode == 0, result.stderr
    assert "Run the NDE narratives workflow" in result.stdout
    assert "Commands:" in result.stdout
    assert "Use 'nde <command> --help' for command-specific help." in result.stdout
    assert "Execute configured LLM experiments and resume" in result.stdout


def test_config_help_shows_clean_repo_relative_defaults() -> None:
    result = run_cli("validate-config", "--help")

    assert result.returncode == 0, result.stderr
    assert "config/study.toml" in result.stdout
    assert "config/paths.local.toml" in result.stdout


def test_help_supports_forced_color_output() -> None:
    result = run_cli("--help", env={"FORCE_COLOR": "1"})

    assert result.returncode == 0, result.stderr
    assert "\x1b[" in result.stdout


def test_run_llm_help_includes_selection_and_examples() -> None:
    result = run_cli("run-llm", "--help")

    assert result.returncode == 0, result.stderr
    assert "Existing artifacts are resumed in place" in result.stdout
    assert "Experiment Selection:" in result.stdout
    assert "Run every enabled [[llm.experiments]] entry" in result.stdout
    assert "Examples:" in result.stdout
    assert "nde run-llm --all-experiments" in result.stdout


def test_local_demo_help_includes_ui_options_and_examples() -> None:
    result = run_cli("local-demo", "--help")

    assert result.returncode == 0, result.stderr
    assert "Start a local Gradio interface" in result.stdout
    assert "--host HOST" in result.stdout
    assert "--port PORT" in result.stdout
    assert "--share" in result.stdout
    assert "nde local-demo --port 7870" in result.stdout


def test_evaluate_help_explains_discovery_and_vader_behavior() -> None:
    result = run_cli("evaluate", "--help")

    assert result.returncode == 0, result.stderr
    assert "LLM artifacts are discovered from llm_results_dir" in result.stdout
    assert "existing VADER scores" in result.stdout
    assert "default VADER score file" in result.stdout
    assert "sample-level" in result.stdout
    assert "automatically" in result.stdout
    assert "Directory to scan for LLM result artifacts" in result.stdout
    assert "Examples:" in result.stdout


def test_build_annotation_sample_creates_workbooks(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    annotation_workbook = tmp_path / "annotation_outputs" / "nde_annotation_sample.xlsx"
    mapping_workbook = (
        tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx"
    )
    column_map_workbook = (
        tmp_path / "annotation_outputs" / "nde_annotation_column_mapping_private.xlsx"
    )

    assert annotation_workbook.exists()
    assert mapping_workbook.exists()
    assert column_map_workbook.exists()

    workbook = load_workbook(annotation_workbook)
    assert set(workbook.sheetnames) == {"annotation", "instructions"}
    worksheet = workbook["annotation"]
    headers = [cell.value for cell in worksheet[1]]
    assert PARTICIPANT_CODE_HEADER in headers
    assert "Context Narrative" in headers
    assert len(worksheet.data_validations.dataValidation) == 2

    sampled_private = pd.read_excel(mapping_workbook, sheet_name="sampled_private")
    assert len(sampled_private) == 2
    assert "participant_code" in sampled_private.columns


def test_build_annotation_sample_refuses_to_overwrite_without_force(
    tmp_path: Path,
) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    first_result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )
    assert first_result.returncode == 0, first_result.stderr

    second_result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert second_result.returncode == 1
    assert "Refusing to overwrite existing annotation artifacts" in second_result.stderr


def test_build_annotation_sample_prefers_preprocessed_dataset_when_available(
    tmp_path: Path,
) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    preprocessed_dir = tmp_path / "preprocessing_outputs"
    preprocessed_dir.mkdir(parents=True, exist_ok=True)
    preprocessed_csv = preprocessed_dir / "cleaned_dataset.csv"

    source = pd.read_csv(survey_csv)
    one_row = source.head(1).copy()
    one_row.loc[:, "response_id"] = 999
    if "m11_quality_label" in one_row.columns:
        one_row = one_row.drop(columns=["m11_quality_label"])
    one_row.to_csv(preprocessed_csv, index=False)

    result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert payload["source_path"].endswith("preprocessing_outputs/cleaned_dataset.csv")

    sampled_private = pd.read_excel(
        tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx",
        sheet_name="sampled_private",
    )
    assert len(sampled_private) == 1
    assert int(sampled_private.loc[0, "response_id"]) == 999


def test_build_annotation_sample_prefers_preprocessed_dataset_over_translated(
    tmp_path: Path,
) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    preprocessed_dir = tmp_path / "preprocessing_outputs"
    preprocessed_dir.mkdir(parents=True, exist_ok=True)

    cleaned = pd.read_csv(survey_csv).head(1).copy()
    cleaned.loc[:, "response_id"] = 901
    cleaned.to_csv(preprocessed_dir / "cleaned_dataset.csv", index=False)

    translated = pd.read_csv(survey_csv).head(1).copy()
    translated.loc[:, "response_id"] = 902
    translated.to_csv(preprocessed_dir / "translated_dataset.csv", index=False)

    result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert payload["source_path"].endswith("preprocessing_outputs/cleaned_dataset.csv")

    sampled_private = pd.read_excel(
        tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx",
        sheet_name="sampled_private",
    )
    assert len(sampled_private) == 1
    assert int(sampled_private.loc[0, "response_id"]) == 901


def test_build_annotation_sample_falls_back_to_survey_when_preprocessed_missing_valence(
    tmp_path: Path,
) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    preprocessed_dir = tmp_path / "preprocessing_outputs"
    preprocessed_dir.mkdir(parents=True, exist_ok=True)
    preprocessed_csv = preprocessed_dir / "cleaned_dataset.csv"

    source = pd.read_csv(survey_csv)
    broken = source.head(1).copy().drop(columns=["valence"])
    broken.to_csv(preprocessed_csv, index=False)

    result = run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    payload = json.loads(result.stdout)
    assert payload["source_path"].endswith("survey_fixture.csv")

    sampled_private = pd.read_excel(
        tmp_path / "annotation_outputs" / "nde_annotation_mapping_private.xlsx",
        sheet_name="sampled_private",
    )
    assert len(sampled_private) == 2


def test_build_llm_batch_writes_experiment_directory_and_manifest(
    tmp_path: Path,
) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    paths_config = make_paths_config(tmp_path, survey_csv)

    run_cli(
        "build-annotation-sample",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )
    result = run_cli(
        "build-llm-batch",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
        "--experiment-id",
        "exp-alpha",
        "--prompt-variant",
        "baseline",
        "--run-id",
        "run-01",
    )

    assert result.returncode == 0, result.stderr

    batch_dir = tmp_path / "llm_batches" / "exp-alpha__run-01"
    manifest_path = batch_dir / "manifest.json"
    assert manifest_path.exists()
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["experiment_id"] == "exp-alpha"
    assert manifest["prompt_variant"] == "baseline"
    assert manifest["run_id"] == "run-01"

    for section in ("context", "experience", "aftereffects"):
        batch_path = batch_dir / f"{section}_batch.jsonl"
        assert batch_path.exists()
        records = [
            json.loads(line)
            for line in batch_path.read_text(encoding="utf-8").splitlines()
            if line
        ]
        assert len(records) == 3
        assert {record["section"] for record in records} == {section}
        assert all(
            record["experiment"]["artifact_id"] == "exp-alpha__run-01"
            for record in records
        )


def test_validate_config_accepts_minimal_data_dir_paths_config(tmp_path: Path) -> None:
    study_config = FIXTURES / "study_test.toml"
    survey_csv = FIXTURES / "survey_fixture.csv"
    data_dir = tmp_path / "nde_data"
    data_dir.mkdir(parents=True, exist_ok=True)
    content = f"""[paths]
data_dir = "{data_dir.as_posix()}"
survey_csv = "{survey_csv.as_posix()}"
"""
    paths_config = tmp_path / "paths.local.toml"
    paths_config.write_text(content, encoding="utf-8")

    result = run_cli(
        "validate-config",
        "--study-config",
        str(study_config),
        "--paths-config",
        str(paths_config),
    )

    assert result.returncode == 0, result.stderr
    assert (data_dir / "annotation_outputs").exists()
    assert (data_dir / "human_annotations").exists()
    assert (data_dir / "llm_outputs").exists()
    assert (data_dir / "preprocessing_outputs").exists()


def test_load_benchmark_config_parses_multiple_datasets(tmp_path: Path) -> None:
    config_path = tmp_path / "paths.local.toml"
    config_path.write_text(
        """[paths]
data_dir = "/tmp/nde-data"
survey_csv = "/tmp/nde-data/survey.csv"

[benchmark.runtime]
provider = "ollama"
base_url = "http://localhost:11434"
timeout_seconds = 120
max_attempts = 2
temperature = 0.0

[benchmark.dataset]
dataset_name = "amazon_reviews_multi"
dataset_config = "en"
split = "train"
text_column = "review_body"
label_column = "stars"
max_rows = 2000
random_state = 20

[[benchmark.datasets]]
dataset_name = "amazon_reviews_multi"
dataset_config = "en"
split = "train"
text_column = "review_body"
label_column = "stars"
max_rows = 100
random_state = 20

[[benchmark.datasets]]
dataset_name = "imdb"
dataset_config = ""
split = "test"
text_column = "text"
label_column = "label"
max_rows = 200
random_state = 20
""",
        encoding="utf-8",
    )

    benchmark = load_benchmark_config(config_path)
    assert len(benchmark.datasets) == 2
    assert benchmark.datasets[0].dataset_name == "amazon_reviews_multi"
    assert benchmark.datasets[1].dataset_name == "imdb"


def test_load_benchmark_config_deduplicates_by_model_and_temperature(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "paths.local.toml"
    config_path.write_text(
        """[paths]
data_dir = "/tmp/nde-data"
survey_csv = "/tmp/nde-data/survey.csv"

[llm]
provider = "ollama"
base_url = "http://localhost:11434"
timeout_seconds = 120
max_attempts = 2
temperature = 0.0

[benchmark.runtime]
provider = "ollama"
base_url = "http://localhost:11434"
timeout_seconds = 120
max_attempts = 2
temperature = 0.0

[[llm.experiments]]
experiment_id = "exp_a"
model = "qwen3.5:9b"
run_id = "01"
temperature = 0.0

[[llm.experiments]]
experiment_id = "exp_b"
model = "qwen3.5:9b"
run_id = "RA1"
temperature = 0.0

[[llm.experiments]]
experiment_id = "exp_c"
model = "qwen3.5:9b"
run_id = "T01"
temperature = 0.2

[[llm.experiments]]
experiment_id = "exp_d"
model = "llama3.1:8b"
run_id = "01"
""",
        encoding="utf-8",
    )

    benchmark = load_benchmark_config(config_path)
    deduped_pairs = {
        (experiment.model, experiment.temperature)
        for experiment in benchmark.experiments
    }

    assert len(benchmark.experiments) == 3
    assert deduped_pairs == {
        ("qwen3.5:9b", 0.0),
        ("qwen3.5:9b", 0.2),
        ("llama3.1:8b", None),
    }
